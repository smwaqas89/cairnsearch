"""Enhanced Excel processor with sheet profiling and intelligent extraction."""
import json
import logging
import re
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from cairnsearch.config import get_config
from cairnsearch.core.models import (
    ProcessingResult, ExtractionMetadata, ChunkMetadata, GuardrailLimits,
)
from cairnsearch.core.exceptions import ProcessingError, GuardrailExceeded
from cairnsearch.core.guardrails import GuardrailEnforcer


logger = logging.getLogger(__name__)


@dataclass
class ColumnInfo:
    """Information about a column."""
    name: str
    index: int
    dtype: str  # text, number, date, currency, percentage, boolean
    sample_values: List[Any]
    non_empty_count: int
    unique_count: int
    is_semantic: bool = True  # Should be included in embeddings
    is_numeric_only: bool = False  # Pure numbers, exclude from embeddings


@dataclass  
class TableRegion:
    """A detected table region within a sheet."""
    start_row: int
    end_row: int
    start_col: int
    end_col: int
    header_row: int
    columns: List[ColumnInfo]
    row_count: int


@dataclass
class SheetProfile:
    """Profile of an Excel sheet."""
    name: str
    index: int
    is_hidden: bool
    is_empty: bool
    row_count: int
    col_count: int
    tables: List[TableRegion]
    has_merged_cells: bool
    header_rows: List[int]
    skip_reason: Optional[str] = None


@dataclass
class RowSummary:
    """Summary of a row for embedding."""
    sheet_name: str
    row_number: int
    summary_text: str
    metadata: Dict[str, Any]
    semantic_columns: List[str]
    numeric_columns: Dict[str, Any]


class EnhancedExcelProcessor:
    """
    Enhanced Excel processor with:
    - Workbook profiling
    - Hidden/empty sheet detection
    - Multiple table detection per sheet
    - Column type inference
    - Merged cell handling
    - Row-level summaries for embedding
    - Numeric column separation
    - Smart guardrail enforcement
    """
    
    # Column type patterns
    DATE_PATTERNS = [
        r'\d{4}-\d{2}-\d{2}',
        r'\d{2}/\d{2}/\d{4}',
        r'\d{2}-\d{2}-\d{4}',
    ]
    CURRENCY_PATTERN = r'^[$€£¥]\s*[\d,]+\.?\d*$|^[\d,]+\.?\d*\s*[$€£¥]$'
    PERCENTAGE_PATTERN = r'^[\d.]+\s*%$'
    
    def __init__(
        self,
        limits: Optional[GuardrailLimits] = None,
    ):
        self.config = get_config()
        self.limits = limits or GuardrailLimits()
        self.guardrails = GuardrailEnforcer(self.limits)
    
    def process(self, file_path: Path) -> ProcessingResult:
        """
        Process an Excel file with enhanced extraction.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            ProcessingResult with sheets, text, and row summaries
        """
        start_time = time.time()
        file_path = Path(file_path)
        
        self.guardrails.start_processing()
        
        try:
            # Check file size
            self.guardrails.enforce(
                self.guardrails.check_file_size(file_path.stat().st_size)
            )
            
            return self._process_excel_internal(file_path)
            
        except GuardrailExceeded as e:
            return ProcessingResult(
                success=False,
                error=str(e),
                error_stage="guardrails",
                processing_time_ms=(time.time() - start_time) * 1000,
            )
        except Exception as e:
            logger.exception(f"Excel processing failed: {e}")
            return ProcessingResult(
                success=False,
                error=str(e),
                error_stage="extraction",
                processing_time_ms=(time.time() - start_time) * 1000,
            )
    
    def _process_excel_internal(self, file_path: Path) -> ProcessingResult:
        """Internal Excel processing."""
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
        
        start_time = time.time()
        
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
        except InvalidFileException:
            return ProcessingResult(
                success=False,
                error="Invalid or corrupted Excel file",
                error_stage="open",
                processing_time_ms=(time.time() - start_time) * 1000,
            )
        
        # Check sheet count
        self.guardrails.enforce(
            self.guardrails.check_sheet_count(len(wb.sheetnames))
        )
        
        # Profile all sheets
        sheet_profiles: List[SheetProfile] = []
        for idx, sheet_name in enumerate(wb.sheetnames):
            profile = self._profile_sheet(wb[sheet_name], sheet_name, idx)
            sheet_profiles.append(profile)
        
        # Process sheets
        all_text_parts = []
        row_summaries: List[RowSummary] = []
        warnings = []
        total_rows = 0
        
        for profile in sheet_profiles:
            if profile.skip_reason:
                warnings.append(f"Sheet '{profile.name}' skipped: {profile.skip_reason}")
                continue
            
            if profile.is_hidden:
                warnings.append(f"Sheet '{profile.name}' is hidden, skipping")
                continue
            
            if profile.is_empty:
                continue
            
            # Check row limit
            if profile.row_count > self.limits.max_rows_per_sheet:
                warnings.append(
                    f"Sheet '{profile.name}' has {profile.row_count} rows, "
                    f"truncating to {self.limits.max_rows_per_sheet}"
                )
            
            # Extract sheet content
            sheet_text, sheet_summaries = self._extract_sheet(
                wb[profile.name],
                profile,
            )
            
            all_text_parts.append(f"=== Sheet: {profile.name} ===")
            all_text_parts.append(sheet_text)
            row_summaries.extend(sheet_summaries)
            total_rows += min(profile.row_count, self.limits.max_rows_per_sheet)
        
        wb.close()
        
        # Combine text
        full_text = "\n\n".join(all_text_parts)
        
        # Build metadata
        metadata = ExtractionMetadata(
            file_path=str(file_path),
            filename=file_path.name,
            file_type="xlsx",
            page_count=len([p for p in sheet_profiles if not p.skip_reason]),
            extraction_method="direct",
            processing_time_ms=(time.time() - start_time) * 1000,
            total_chars=len(full_text),
            has_tables=True,
            warnings=warnings,
        )
        
        return ProcessingResult(
            success=True,
            text=full_text,
            chunks=[],  # Row summaries will be chunked separately
            metadata=metadata,
            warnings=warnings,
            processing_time_ms=(time.time() - start_time) * 1000,
        )
    
    def _profile_sheet(
        self,
        sheet,
        sheet_name: str,
        sheet_index: int,
    ) -> SheetProfile:
        """Profile an Excel sheet."""
        # Check if hidden
        is_hidden = sheet.sheet_state == 'hidden'
        
        # Get dimensions
        try:
            max_row = sheet.max_row or 0
            max_col = sheet.max_column or 0
        except:
            max_row = 0
            max_col = 0
        
        is_empty = max_row == 0 or max_col == 0
        
        # Check for extremely wide sheets
        skip_reason = None
        if max_col > 100:
            skip_reason = f"Too many columns ({max_col})"
        elif max_row > self.limits.max_rows_per_sheet * 2:
            skip_reason = f"Too many rows ({max_row})"
        
        if is_empty or skip_reason:
            return SheetProfile(
                name=sheet_name,
                index=sheet_index,
                is_hidden=is_hidden,
                is_empty=is_empty,
                row_count=max_row,
                col_count=max_col,
                tables=[],
                has_merged_cells=False,
                header_rows=[],
                skip_reason=skip_reason,
            )
        
        # Detect merged cells
        has_merged_cells = bool(sheet.merged_cells.ranges) if hasattr(sheet, 'merged_cells') else False
        
        # Detect header rows and tables
        header_rows = self._detect_header_rows(sheet, max_row, max_col)
        tables = self._detect_table_regions(sheet, header_rows, max_row, max_col)
        
        return SheetProfile(
            name=sheet_name,
            index=sheet_index,
            is_hidden=is_hidden,
            is_empty=is_empty,
            row_count=max_row,
            col_count=max_col,
            tables=tables,
            has_merged_cells=has_merged_cells,
            header_rows=header_rows,
        )
    
    def _detect_header_rows(
        self,
        sheet,
        max_row: int,
        max_col: int,
    ) -> List[int]:
        """Detect potential header rows in a sheet."""
        header_rows = []
        
        # Check first 10 rows for header patterns
        for row_idx in range(1, min(11, max_row + 1)):
            row_values = []
            for col_idx in range(1, min(max_col + 1, 50)):
                try:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        row_values.append(str(cell.value))
                except:
                    pass
            
            if not row_values:
                continue
            
            # Header heuristics
            is_header = True
            
            # Check if all values are strings
            all_strings = all(isinstance(v, str) for v in row_values)
            
            # Check if values look like column names
            name_like = sum(1 for v in row_values if self._is_column_name_like(v))
            
            # Check if most values are unique
            unique_ratio = len(set(row_values)) / len(row_values) if row_values else 0
            
            if all_strings and name_like > len(row_values) * 0.5 and unique_ratio > 0.8:
                header_rows.append(row_idx)
        
        return header_rows[:3]  # Return at most 3 header rows
    
    def _is_column_name_like(self, value: str) -> bool:
        """Check if a value looks like a column name."""
        if not value or len(value) > 50:
            return False
        
        # Contains letters
        if not any(c.isalpha() for c in value):
            return False
        
        # Not too many numbers
        num_count = sum(1 for c in value if c.isdigit())
        if num_count > len(value) * 0.5:
            return False
        
        # Common column name patterns
        patterns = [
            r'^[A-Z][a-z]+',  # Capitalized words
            r'^[A-Z_]+$',    # ALL CAPS
            r'\w+\s\w+',     # Multi-word
        ]
        
        return any(re.match(p, value) for p in patterns)
    
    def _detect_table_regions(
        self,
        sheet,
        header_rows: List[int],
        max_row: int,
        max_col: int,
    ) -> List[TableRegion]:
        """Detect distinct table regions in a sheet."""
        tables = []
        
        if not header_rows:
            # Assume single table starting at row 1
            columns = self._analyze_columns(sheet, 1, max_row, max_col)
            tables.append(TableRegion(
                start_row=1,
                end_row=max_row,
                start_col=1,
                end_col=max_col,
                header_row=1,
                columns=columns,
                row_count=max_row,
            ))
        else:
            # Create table region for each header row
            for i, header_row in enumerate(header_rows):
                # Find end of this table
                if i < len(header_rows) - 1:
                    end_row = header_rows[i + 1] - 1
                else:
                    end_row = max_row
                
                columns = self._analyze_columns(sheet, header_row, end_row, max_col)
                
                tables.append(TableRegion(
                    start_row=header_row,
                    end_row=end_row,
                    start_col=1,
                    end_col=max_col,
                    header_row=header_row,
                    columns=columns,
                    row_count=end_row - header_row,
                ))
        
        return tables
    
    def _analyze_columns(
        self,
        sheet,
        start_row: int,
        end_row: int,
        max_col: int,
    ) -> List[ColumnInfo]:
        """Analyze column types and properties."""
        columns = []
        sample_rows = min(100, end_row - start_row)
        
        for col_idx in range(1, min(max_col + 1, 100)):
            # Get header
            header_cell = sheet.cell(row=start_row, column=col_idx)
            header = str(header_cell.value) if header_cell.value else f"Column_{col_idx}"
            
            # Sample values
            values = []
            for row_idx in range(start_row + 1, min(start_row + sample_rows + 1, end_row + 1)):
                try:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        values.append(cell.value)
                except:
                    pass
            
            # Infer type
            dtype = self._infer_column_type(values)
            
            # Determine if semantic or numeric
            is_numeric_only = dtype in ['number', 'currency', 'percentage']
            is_semantic = not is_numeric_only or dtype == 'date'
            
            columns.append(ColumnInfo(
                name=header,
                index=col_idx,
                dtype=dtype,
                sample_values=values[:5],
                non_empty_count=len(values),
                unique_count=len(set(str(v) for v in values)),
                is_semantic=is_semantic,
                is_numeric_only=is_numeric_only,
            ))
        
        return columns
    
    def _infer_column_type(self, values: List[Any]) -> str:
        """Infer the type of a column from sample values."""
        if not values:
            return "text"
        
        type_counts = {
            "number": 0,
            "date": 0,
            "currency": 0,
            "percentage": 0,
            "boolean": 0,
            "text": 0,
        }
        
        for value in values:
            if isinstance(value, bool):
                type_counts["boolean"] += 1
            elif isinstance(value, (int, float)):
                type_counts["number"] += 1
            elif isinstance(value, datetime):
                type_counts["date"] += 1
            elif isinstance(value, str):
                value_str = value.strip()
                
                # Check patterns
                if re.match(self.CURRENCY_PATTERN, value_str):
                    type_counts["currency"] += 1
                elif re.match(self.PERCENTAGE_PATTERN, value_str):
                    type_counts["percentage"] += 1
                elif any(re.match(p, value_str) for p in self.DATE_PATTERNS):
                    type_counts["date"] += 1
                elif value_str.lower() in ('true', 'false', 'yes', 'no'):
                    type_counts["boolean"] += 1
                else:
                    try:
                        float(value_str.replace(',', ''))
                        type_counts["number"] += 1
                    except:
                        type_counts["text"] += 1
            else:
                type_counts["text"] += 1
        
        # Return most common type
        return max(type_counts, key=type_counts.get)
    
    def _extract_sheet(
        self,
        sheet,
        profile: SheetProfile,
    ) -> Tuple[str, List[RowSummary]]:
        """Extract content from a sheet."""
        text_parts = []
        row_summaries = []
        
        max_rows = min(profile.row_count, self.limits.max_rows_per_sheet)
        
        for table in profile.tables:
            # Get semantic and numeric columns
            semantic_cols = [c for c in table.columns if c.is_semantic]
            numeric_cols = [c for c in table.columns if c.is_numeric_only]
            
            # Build header row text
            header_values = []
            for col in table.columns:
                header_values.append(col.name)
            text_parts.append(" | ".join(header_values))
            
            # Process data rows
            for row_idx in range(table.header_row + 1, min(table.end_row + 1, table.header_row + max_rows + 1)):
                row_values = []
                semantic_values = {}
                numeric_values = {}
                
                for col in table.columns:
                    try:
                        cell = sheet.cell(row=row_idx, column=col.index)
                        value = cell.value
                        str_value = str(value) if value is not None else ""
                        row_values.append(str_value)
                        
                        if col.is_semantic and value is not None:
                            semantic_values[col.name] = str_value
                        if col.is_numeric_only and value is not None:
                            numeric_values[col.name] = value
                    except:
                        row_values.append("")
                
                # Add to text
                if any(row_values):
                    text_parts.append(" | ".join(row_values))
                
                # Create row summary for embedding
                if semantic_values:
                    summary = self._build_row_summary(
                        profile.name,
                        row_idx,
                        semantic_values,
                        numeric_values,
                        [c.name for c in semantic_cols],
                    )
                    row_summaries.append(summary)
        
        return "\n".join(text_parts), row_summaries
    
    def _build_row_summary(
        self,
        sheet_name: str,
        row_number: int,
        semantic_values: Dict[str, str],
        numeric_values: Dict[str, Any],
        semantic_columns: List[str],
    ) -> RowSummary:
        """Build a row summary for embedding."""
        # Build natural language summary
        parts = []
        for key, value in semantic_values.items():
            if value:
                parts.append(f"{key}: {value}")
        
        summary_text = "; ".join(parts)
        
        return RowSummary(
            sheet_name=sheet_name,
            row_number=row_number,
            summary_text=summary_text,
            metadata={
                "sheet": sheet_name,
                "row": row_number,
                **numeric_values,
            },
            semantic_columns=semantic_columns,
            numeric_columns=numeric_values,
        )
    
    def get_row_summaries_for_embedding(
        self,
        file_path: Path,
    ) -> List[RowSummary]:
        """
        Get row summaries specifically for embedding.
        
        This extracts only the semantic content, excluding
        pure numeric columns from embeddings.
        """
        result = self.process(file_path)
        
        if not result.success:
            return []
        
        # Re-process to get summaries (in full implementation, 
        # this would be cached from process())
        return []  # Placeholder
