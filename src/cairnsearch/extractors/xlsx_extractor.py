"""Excel/spreadsheet text extraction."""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .base import BaseExtractor, ExtractionResult


class XlsxExtractor(BaseExtractor):
    """Extract text from Excel files."""

    @property
    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".xlsm", ".xls"]

    def extract(self, file_path: Path) -> ExtractionResult:
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            all_text = []
            
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                sheet_text = [f"=== Sheet: {sheet_name} ==="]
                
                for row in sheet.iter_rows():
                    row_values = []
                    for cell in row:
                        if cell.value is not None:
                            row_values.append(str(cell.value))
                    if row_values:
                        sheet_text.append(" | ".join(row_values))
                
                all_text.extend(sheet_text)
            
            wb.close()
            text = "\n".join(all_text)

            # Excel files don't have standard metadata
            return ExtractionResult(
                success=True,
                text=text,
                metadata={
                    "page_count": len(wb.sheetnames),
                },
                extraction_method="direct"
            )
        except InvalidFileException:
            return ExtractionResult(
                success=False,
                error="Invalid or corrupted Excel file"
            )
        except Exception as e:
            return ExtractionResult(success=False, error=str(e))
